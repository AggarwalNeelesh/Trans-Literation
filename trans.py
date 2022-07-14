from colorama import init
import openpyxl
from googletrans import Translator


def transliteration(msg):
    # Method for translating any language into English
    translator = Translator()
    data = translator.translate(msg).text
    return data


init(convert=True)
path = "C:\\tranlitrated_test.xlsx"

wb_obj = openpyxl.load_workbook(path.strip())
sheet_obj = wb_obj.active

# get max row count
max_row = sheet_obj.max_row

for j in range(2, max_row):
    word = sheet_obj.cell(row=j, column=1)
    if word.value is None:
        break
    msg = transliteration(word.value)
    sheet_obj.cell(row=j, column=2).value = msg
    print(word.value, " ", msg)
    wb_obj.save("C:\\tranlitrated_test.xlsx")
